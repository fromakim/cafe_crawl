# %% In[0]: Basic Settings
baseURL = 'https://cafe.naver.com/cfcmania/ArticleList.nhn?search.clubid=25204640&search.menuid=29&search.boardtype=L&search.specialmenutype=&userDisplay=50'
baseCafeURL = 'https://cafe.naver.com/cfcmania/'
pageMax = 1

# %% In[1]: import packages
from selenium import webdriver
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook

# %% In[2]: Driver Setting
driver = webdriver.Chrome('./../chromedriver')
driver.implicitly_wait(10)

workbook = Workbook()
workbook.create_sheet('articles')
workbook.create_sheet('comments')
workbook.remove(workbook['Sheet'])
workbook.save('data.xlsx')

ar = workbook['articles']
ar.append(['ID', 'Title', 'Author', 'Text', 'Image'])

co = workbook['comments']
row = ['ID', 'ArticleAuthor', 'TopLevelCommentAuthor', 'TopLevelCommentText']
for i in range(10):
    row.append(str(i + 1) + '_CommentAuthor')
    row.append(str(i + 1) + '_CommentObject')
    row.append(str(i + 1) + '_CommentText')
co.append(row)

workbook.save('data.xlsx')

# %% In[3]: Login
driver.get('https://www.naver.com')
login = driver.find_element_by_class_name('lg_local_btn')
login.click()
# TODO: 로그인은 그냥 수동으로, 필요할 시 생략할 것
input('Login and then Press Enter: ')

# %% In[4]: Retrieve Article List
article = []
for i in range(pageMax):
    URL = baseURL + '&search.page=' + str(i + 1)

    driver.get(URL)

    time.sleep(1)

    iframes = driver.find_elements_by_tag_name('iframe')
    soup = None

    for i, iframe in enumerate(iframes):
        if iframe.get_attribute('title') == '카페 메인':
            driver.switch_to.frame(iframes[i])

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            driver.switch_to.default_content()

    for s in soup.select('div.inner_number'):
        article.append(s.text)

# %% In[5]: Get Pages
pages = []
for articleNum in article:
    print('[%d/%d] Retrieving Pages' % (article.index(articleNum) + 1, len(article)))
    URL = baseCafeURL + articleNum
    driver.get(URL)
    time.sleep(1)

    iframes = driver.find_elements_by_tag_name('iframe')
    soup = None

    for i, iframe in enumerate(iframes):
        if iframe.get_attribute('title') == '카페 메인':
            driver.switch_to.frame(iframes[i])

            soup = BeautifulSoup(driver.page_source, 'html.parser')

            driver.switch_to.default_content()

    pages.append(soup)
driver.close()

# %% In[6]: Get Article Info
for page in pages:
    print('[%d/%d] Retrieving Articles' % (pages.index(page) + 1, len(pages)))

    title = page.select_one('td > span.b.m-tcol-c:not(.reply)').text
    articleID = page.select_one('a#linkUrl').text.split('/')[-1]
    author = page.select_one('td.p-nick > a.m-tcol-c.b').text.split('(')[0]
    text = ''.join(list(map(lambda a : a.text + '\n', page.select('div.tbody.m-tcol-c#tbody :not(img):not(br)'))))
    icode = ''.join(list(map(lambda a : a.attrs['src'] + '\n\n', page.select('div.tbody.m-tcol-c#tbody img'))))

    ar.append([articleID, title, author, text, icode])
workbook.save('data.xlsx')

# %% In[7]: Get Comments Info
for page in pages:
    print('[%d/%d] Retrieving Comments' % (pages.index(page) + 1, len(pages)))

    articleID = page.select_one('a#linkUrl').text.split('/')[-1]
    author = page.select_one('td.p-nick > a.m-tcol-c.b').text.split('(')[0]

    topLevelComments = page.select('#cmt_list > li:not(.reply):not(.filter-30)')
    allComments = page.select('#cmt_list > li:not(.filter-30)')

    for tlc in topLevelComments:
        cauthor = tlc.select_one('a._nickUI').text
        ctext = tlc.select_one('span.comm_body').text

        index = allComments.index(tlc)
        nindex = topLevelComments.index(tlc) + 1

        window = []
        replies = []

        window = allComments[index:allComments.index(topLevelComments[nindex])] if nindex < len(topLevelComments) else allComments[index:]

        replies = list(map(lambda e : [e.select_one('a.m-tcol-c._rosRestrict._nickUI').text, (e.select_one('a.m-tcol-c.filter-50.nick').text) if e.select_one('a.m-tcol-c.filter-50.nick') else None, e.select_one('span.comm_body').text], window))
        replies = [item for sublist in replies for item in sublist]

        co.append([articleID, author, cauthor, ctext, *replies])
workbook.save('data.xlsx')
